import openpyxl
from fpdf import FPDF

def DatuEksportesana(rekinu_saraksts, fails, formats="xlsx"):
    """
    Pieņem list rekinu_saraksts, str fails un str formats, atgriež str fails.
    Eksportē datus uz .xlsx vai .pdf failu.
    """
    if not rekinu_saraksts:
        raise ValueError("Kļūda: rēķinu saraksts ir tukšs.")
    if formats == "xlsx":
        return _uz_xlsx(rekinu_saraksts, fails)
    elif formats == "pdf":
        return _uz_pdf(rekinu_saraksts, fails)
    else:
        raise ValueError(f"Kļūda: nezināms formāts '{formats}'.")


def _rindas(rekinu_saraksts):
    """Palīgfunkcija - izveido rindu sarakstu no rēķinu datiem."""
    return [
        (rekins.get("periods", "?"), p.get("veids", "?"),
         p["paterins"], p["tarifs"], round(p["paterins"] * p["tarifs"], 2))
        for rekins in rekinu_saraksts
        for p in rekins.get("pakalpojumi", [])
    ]


def _uz_xlsx(rekinu_saraksts, fails):
    """Eksportē uz Excel failu."""
    if not fails.endswith(".xlsx"):
        fails += ".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rēķini"
    ws.append(["Periods", "Pakalpojums", "Patēriņš", "Tarifs", "Summa (EUR)"])
    for rinda in _rindas(rekinu_saraksts):
        ws.append(list(rinda))
    wb.save(fails)
    return fails


def _uz_pdf(rekinu_saraksts, fails):
    """Eksportē uz PDF failu."""
    if not fails.endswith(".pdf"):
        fails += ".pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(200, 10, "Komunālo pakalpojumu rēķini", ln=True, align="C")
    pdf.ln(3)

    # Galvene
    pdf.set_font("Arial", style="B", size=10)
    for virsraksts, platums in [("Periods", 35), ("Pakalpojums", 50), ("Patēriņš", 30), ("Tarifs", 30), ("Summa", 30)]:
        pdf.cell(platums, 8, virsraksts, border=1)
    pdf.ln()

    # Dati
    pdf.set_font("Arial", size=10)
    for periods, veids, paterins, tarifs, summa in _rindas(rekinu_saraksts):
        for vertiba, platums in [(periods, 35), (veids, 50), (paterins, 30), (tarifs, 30), (summa, 30)]:
            pdf.cell(platums, 8, str(vertiba), border=1)
        pdf.ln()

    pdf.output(fails)
    return fails
