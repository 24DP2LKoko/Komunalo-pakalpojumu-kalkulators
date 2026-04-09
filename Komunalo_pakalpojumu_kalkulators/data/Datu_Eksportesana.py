import openpyxl
from fpdf import FPDF


def DatuEksportesana(rekinu_saraksts, fails, formats="xlsx"):
    """
    funkcija DatuEksportesana pieņem list tipa vērtību rekinu_saraksts, str tipa vērtību fails un str tipa vērtību formats un atgriež str tipa vērtību fails
    """

    if not rekinu_saraksts:
        raise ValueError("Kļūda: rēķinu saraksts ir tukšs.")

    rindas = []

    for rekins in rekinu_saraksts:
        rindas.append((
            rekins.get("id", ""),
            rekins.get("veids", ""),
            rekins.get("summa", 0),
            rekins.get("datums", ""),
            rekins.get("periods", "")
        ))

    if formats == "xlsx":
        if not fails.endswith(".xlsx"):
            fails += ".xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["ID", "Veids", "Summa (EUR)", "Datums", "Periods"])

        for rinda in rindas:
            ws.append(list(rinda))

        wb.save(fails)
        return fails

    elif formats == "pdf":
        if not fails.endswith(".pdf"):
            fails += ".pdf"

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=10)

        pdf.cell(20, 10, "ID", 1)
        pdf.cell(40, 10, "Veids", 1)
        pdf.cell(35, 10, "Summa", 1)
        pdf.cell(45, 10, "Datums", 1)
        pdf.cell(40, 10, "Periods", 1)
        pdf.ln()

        for rinda in rindas:
            pdf.cell(20, 10, str(rinda[0]), 1)
            pdf.cell(40, 10, str(rinda[1]), 1)
            pdf.cell(35, 10, str(rinda[2]), 1)
            pdf.cell(45, 10, str(rinda[3]), 1)
            pdf.cell(40, 10, str(rinda[4]), 1)
            pdf.ln()

        pdf.output(fails)
        return fails

    else:
        raise ValueError(f"Kļūda: nezināms formāts '{formats}'.")